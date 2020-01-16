#!/usr/bin/env python
# coding: utf-8

# In[21]:


import numpy as np
import tensorflow as tf
import math
import matplotlib.pyplot as plt
from tensorflow.python.framework import ops
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import ot.plot
from sklearn.metrics import mean_squared_error
from math import sqrt
import ot
class TD_Network(object):
    def __init__(self):
        self.X = tf.placeholder(tf.float32, [None,15], name='Placeholder_X')
        self.Y = tf.placeholder(tf.float32, [None,1], name='Placeholder_Y')
        
        with tf.variable_scope('TDANN'):
            self.Ypred = self.TD_ANN(self.X)
        
        self.lr = 1e-4
        self.global_step = tf.Variable(0, trainable = False)
        self.decay = tf.train.exponential_decay(self.lr, self.global_step, 100, 0.99, staircase = True)
        self.loss = tf.reduce_mean(tf.losses.mean_squared_error(self.Y, self.Ypred))
        self.optimizer = tf.train.AdamOptimizer(self.decay, name='Adam').minimize(self.loss)
        
        
    def TD_ANN(self,X):
        self.Hidden1 = tf.contrib.layers.fully_connected(X, 120, activation_fn = tf.nn.relu,
                                                         weights_initializer = tf.contrib.layers.xavier_initializer())
        self.Hidden2 = tf.contrib.layers.fully_connected(self.Hidden1, 70, activation_fn = tf.nn.relu,
                                                         weights_initializer = tf.contrib.layers.xavier_initializer())
        self.Hidden3 = tf.contrib.layers.fully_connected(self.Hidden2, 35, activation_fn = tf.nn.relu,
                                                         weights_initializer = tf.contrib.layers.xavier_initializer())
        self.Out = tf.contrib.layers.fully_connected(self.Hidden3, 1, activation_fn = None,
                                                         weights_initializer = tf.contrib.layers.xavier_initializer())
        return self.Out

def preprocessing(data, sequence_length):
    len_data = np.shape(data)[0]
    batchdataindex = range(sequence_length, len_data)
    permindex = np.array(batchdataindex)
    rng = np.random.RandomState(23456)
    rng.shuffle(permindex)
    return permindex

def preprocessing_lin(data, sequence_length):
    len_data = np.shape(data)[0]
    batchdataindex = range(sequence_length, len_data)
    permindex = np.array(batchdataindex)
    return permindex
        
def load_data():
    wb_Tr = Workbook()
    wb_Tr = load_workbook(filename = 'Mass_total.xlsx', data_only = True)
    ws_Tr = wb_Tr.active
    Train_num = 2000

    X_train1 = np.zeros((Train_num,1))
    Y_train1 = np.zeros((Train_num,1))
    X_train2 = np.zeros((Train_num,1))
    Y_train2 = np.zeros((Train_num,1))
    
    for i in range(Train_num):
        X_train1[i][0] = ws_Tr.cell(row=i+1, column=1).value
        Y_train1[i][0] = ws_Tr.cell(row=i+1, column=2).value
        X_train2[i][0] = ws_Tr.cell(row=i+1, column=5).value
        Y_train2[i][0] = ws_Tr.cell(row=i+1, column=6).value
    
    return [X_train1, Y_train1, X_train2, Y_train2]

def train(X_train1, Y_train1):
    conf = tf.ConfigProto(allow_soft_placement=True)
    conf.gpu_options.per_process_gpu_memory_fraction = 0.5
    conf.gpu_options.allow_growth = True
    os.environ['CUDA_VISIBLE_DEVICES'] = '0'
    
    sequence_length = 15
    imsiX = np.array(X_train1[0:1700,:])
    imsiY = np.array(Y_train1[0:1700,:])
    train_seq = preprocessing_lin(X_train1[0:1700,:], sequence_length)
    seq_idxs = np.array([train_seq - n for n in reversed(range(0, sequence_length))]).T
    X_train_tune = np.reshape(imsiX[seq_idxs],[-1, sequence_length, 1])
    Y_train_tune = np.reshape(imsiY[seq_idxs],[-1, sequence_length, 1])
    X_train_tune = np.reshape(X_train_tune, [-1, sequence_length])
    Y_train_tune = np.reshape(Y_train_tune[:, sequence_length-1, :], [-1, 1])
    
    minibatch_size = 128
    tf.reset_default_graph()
    model = TD_Network()
    with tf.Session(config=conf) as sess:
        sess.run(tf.global_variables_initializer())
        sess.run(tf.local_variables_initializer())
        for epoch in range(2001):
            A = 0
            for i in range((int)(np.shape(X_train_tune)[0]/minibatch_size)):
                X_batch = X_train_tune[i*minibatch_size:(i+1)*minibatch_size, :]
                Y_batch = Y_train_tune[i*minibatch_size:(i+1)*minibatch_size, :]
                _,a = sess.run([model.optimizer, model.loss], feed_dict={model.X: X_batch, model.Y: Y_batch})
                A = A+a
            A = A / (int)(np.shape(X_train_tune)[0]/minibatch_size)
            if epoch % 100 == 0:
                print("Training Epoch:",epoch,"Loss:",A)
        ckpt_dir = './ckpt/'
        if not os.path.exists(ckpt_dir):
            os.makedirs(ckpt_dir)
        ckpt_path = ckpt_dir + 'TDANN_AIS' + '.ckpt'
        saver = tf.train.Saver()
        saver.save(sess, ckpt_path)

def test_source(X_train1, Y_train1):
    tf.reset_default_graph()
    model = TD_Network()
    conf = tf.ConfigProto(allow_soft_placement=True)
    conf.gpu_options.per_process_gpu_memory_fraction = 0.5
    conf.gpu_options.allow_growth = True
    os.environ['CUDA_VISIBLE_DEVICES'] = '0'
    sequence_length = 15
    X_test = X_train1[1700:1850,:]
    Y_test = Y_train1[1700:1850,:]

    imsiX = np.array(X_test)
    imsiY = np.array(Y_test)
    test_seq = preprocessing_lin(X_test, sequence_length)
    seq_idxs = np.array([test_seq - n for n in reversed(range(0, sequence_length))]).T
    X_test_tune = np.reshape(imsiX[seq_idxs],[-1, sequence_length, 1])
    Y_test_tune = np.reshape(imsiY[seq_idxs],[-1, sequence_length, 1])
    X_test_tune = np.reshape(X_test_tune, [-1, sequence_length])
    Y_test_tune = np.reshape(Y_test_tune[:, sequence_length-1, :], [-1, 1])
    X_input = X_test_tune[:,sequence_length-1]
    
    with tf.Session(config=conf) as sess:
        ckpt_dir = './ckpt/'
        ckpt_path_location = ckpt_dir + 'TDANN_AIS' + '.ckpt'
        saver = tf.train.Saver()
        saver.restore(sess, ckpt_path_location)
        
        ypred = sess.run([model.Ypred], feed_dict={model.X: X_test_tune})
    ypred = np.array(ypred)
    ypred = np.reshape(ypred, [-1])
    #plt.plot(ypred)
    #plt.plot(Y_test_tune)
    #plt.rcParams["figure.figsize"] = (5,2)
    #plt.legend(['Estimation','Reference'])
    rms = sqrt(mean_squared_error(ypred, Y_test_tune))
    print('RMSE:',rms)
    
def OT(X_train1, X_train2):
    sequence_length = 15
    X_target = np.array(X_train2[0:1000,:])
    test_seq = preprocessing_lin(X_target, sequence_length)
    seq_idxs = np.array([test_seq - n for n in reversed(range(0, sequence_length))]).T
    X_target_tune = np.reshape(X_target[seq_idxs],[-1, sequence_length])


    X_source = np.array(X_train1[0:1000,:])
    test_seq = preprocessing_lin(X_source, sequence_length)
    seq_idxs = np.array([test_seq - n for n in reversed(range(0, sequence_length))]).T
    X_source_tune = np.reshape(X_source[seq_idxs],[-1, sequence_length])

    M = ot.dist(X_target_tune, X_source_tune, metric='sqeuclidean')
    ot_emd = ot.da.EMDTransport()
    ot_emd.fit(Xs=X_target_tune, Xt=X_source_tune)

    
    return ot_emd
    
def test_target(X_train2, Y_train2, ot_emd):
    conf = tf.ConfigProto(allow_soft_placement=True)
    conf.gpu_options.per_process_gpu_memory_fraction = 0.5
    conf.gpu_options.allow_growth = True
    os.environ['CUDA_VISIBLE_DEVICES'] = '0'
    tf.reset_default_graph()
    model = TD_Network()
    sequence_length = 15
    X_test = X_train2[0:150,:]
    Y_test = Y_test = Y_train2[0:150,:]
    imsiX = np.array(X_test)
    imsiY= np.array(Y_test)
    test_seq = preprocessing_lin(X_test, sequence_length)
    seq_idxs = np.array([test_seq - n for n in reversed(range(0, sequence_length))]).T
    X_test_tune_1 = np.reshape(imsiX[seq_idxs],[-1, sequence_length])
    X_test_tune_2 = ot_emd.transform(Xs=X_test_tune_1)
    Y_test_tune = np.reshape(imsiY[seq_idxs],[-1, sequence_length, 1])
    X_test_tune = np.reshape(X_test_tune_2, [-1, sequence_length])
    Y_test_tune = np.reshape(Y_test_tune[:, sequence_length-1, :], [-1, 1])
    X_input = X_test_tune[:,sequence_length-1]
    
    with tf.Session(config=conf) as sess:
        ckpt_dir = './ckpt/'
        ckpt_path_location = ckpt_dir + 'TDANN_AIS' + '.ckpt'
        saver = tf.train.Saver()
        saver.restore(sess, ckpt_path_location)
        ypred = sess.run([model.Ypred], feed_dict={model.X: X_test_tune})
        
    ypred = np.array(ypred)
    ypred = np.reshape(ypred, [-1])
    plt.plot(ypred)
    plt.plot(Y_test_tune)
    plt.rcParams["figure.figsize"] = (5,2)
    plt.legend(['Estimation','Reference'])
    rms = sqrt(mean_squared_error(ypred, Y_test_tune))
    print('RMSE:',rms)
    
if __name__ == '__main__':
    X_train1, Y_train1, X_train2, Y_train2 = load_data()
    train(X_train1, Y_train1)
    test_source(X_train1, Y_train1)
    otmatrix = OT(X_train1, X_train2)
    test_target(X_train2, Y_train2, otmatrix)


# In[ ]:




